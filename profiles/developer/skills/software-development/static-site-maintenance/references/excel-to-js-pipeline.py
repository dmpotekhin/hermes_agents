"""Concrete example: extracting book data from Excel and generating a JS data file.

Pattern used in dmpotekhin.github.io/books.html — 521 books from Книги.xlsx.
"""
import openpyxl, json, re

# === STEP 1: Extract from Excel ===
wb = openpyxl.load_workbook("Книги.xlsx")
ws = wb.active

books = []
for r in range(2, ws.max_row + 1):
    author = ws.cell(r, 1).value
    title = ws.cell(r, 2).value
    genre = ws.cell(r, 3).value

    if not author or not title:
        continue

    books.append({
        "author": str(author).strip(),
        "title": str(title).strip(),
        "genre": str(genre).strip() if genre else "Без категории",
    })

# === STEP 2: Clean data ===
for book in books:
    # Remove literal newlines that break JS strings
    book["title"] = book["title"].replace("\n", " ").replace("\r", " ")
    book["author"] = book["author"].replace("\n", " ").replace("\r", " ")
    # Collapse multiple spaces
    book["title"] = re.sub(r" {2,}", " ", book["title"])
    book["author"] = re.sub(r" {2,}", " ", book["author"])

# === STEP 3: Generate JS file ===
lines = []
lines.append("// Auto-generated from Excel source")
lines.append(f"// Total: {len(books)} entries")
lines.append("const dataName = [")

for i, book in enumerate(books):
    # Escape backslashes and double-quotes
    author = book["author"].replace("\\", "\\\\").replace('"', '\\"')
    title = book["title"].replace("\\", "\\\\").replace('"', '\\"')
    genre = book["genre"].replace("\\", "\\\\").replace('"', '\\"')

    comma = "," if i < len(books) - 1 else ""
    lines.append("  {")
    lines.append(f'    "author": "{author}",')
    lines.append(f'    "title": "{title}",')
    lines.append(f'    "genre": "{genre}"')
    lines.append(f"  }}{comma}")

lines.append("];")

with open("js/books-data.js", "w", encoding="utf-8") as f:
    f.write("\n".join(lines))

# === STEP 4: Verify with Node.js ===
import subprocess
r = subprocess.run(
    ['node', '-e',
     'const fs=require("fs");'
     'const code=fs.readFileSync("js/books-data.js","utf8");'
     'new Function(code);'
     'const arr=JSON.parse(code.slice(code.indexOf("["),code.lastIndexOf("]")+1));'
     'console.log(JSON.stringify({count:arr.length,genres:[...new Set(arr.map(b=>b.genre))]}));'
    ], capture_output=True, text=True)

if r.returncode == 0:
    data = json.loads(r.stdout.strip())
    print(f"OK: {data['count']} entries, {len(data['genres'])} genres: {data['genres']}")
else:
    print(f"SYNTAX ERROR: {r.stderr}")
